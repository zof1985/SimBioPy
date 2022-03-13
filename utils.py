# UTILS MODULE


#! IMPORTS


import numpy as np
import openpyxl as xl
import os
import pandas as pd
import time
from datetime import datetime, date
from typing import Tuple


#! CONSTANTS


datetime_format = "%d/%m/%Y-%H:%M:%S"


#! CLASSES


class Participant:
    """
    class containing all the data relevant to a participant.

    Parameters
    ----------
    df: pandas.DataFrame
        a dataframe resulting from the export of Cosmed Omia.
    """

    # class variables
    _name = None
    _surname = None
    _gender = None
    _height = None
    _weight = None
    _birth_date = None

    def __init__(
        self,
        surname: str = None,
        name: str = None,
        gender: str = None,
        height: Tuple[int, float] = None,
        weight: Tuple[int, float] = None,
        age: Tuple[int, float] = None,
        birth_date: date = None,
    ):
        self.setSurname(surname)
        self.setName(name)
        self.setGender(gender)
        self.setHeight(height / 100)
        self.setWeight(weight)
        self.setAge(age)
        self.setBirthDate(birth_date)

    def setSurname(self, surname):
        """
        set the participant surname.

        Parameters
        ----------
        surname: str
            the surname of the participant.
        """
        if surname is not None:
            assert isinstance(surname, str), "'surname' must be a string."
        self._surname = surname

    def setName(self, name):
        """
        set the participant name.

        Parameters
        ----------
        name: str
            the name of the participant.
        """
        if name is not None:
            assert isinstance(name, str), "'name' must be a string."
        self._name = name

    def setGender(self, gender):
        """
        set the participant gender.

        Parameters
        ----------
        gender: str
            the gender of the participant.
        """
        if gender is not None:
            assert isinstance(gender, str), "'gender' must be a string."
        self._gender = gender

    def setHeight(self, height):
        """
        set the participant height in meters.

        Parameters
        ----------
        height: int, float
            the height of the participant.
        """
        if height is not None:
            txt = "'height' must be a float or int."
            assert isinstance(height, (int, float)), txt
        self._height = height

    def setWeight(self, weight):
        """
        set the participant weight in kg.

        Parameters
        ----------
        weight: int, float
            the weight of the participant.
        """
        if weight is not None:
            txt = "'weight' must be a float or int."
            assert isinstance(weight, (int, float)), txt
        self._weight = weight

    def setAge(self, age):
        """
        set the participant age in years.


        Parameters
        ----------
        age: int, float
            the age of the participant.
        """
        if age is not None:
            txt = "'age' must be a float or int."
            assert isinstance(age, (int, float)), txt
        self._age = age

    def setBirthDate(self, birth_date):
        """
        set the participant birth_date.

        Parameters
        ----------
        birth_date: datetime.date
            the birth date of the participant.
        """
        if birth_date is not None:
            txt = "'BirthDate' must be a datetime.date or datetime.datetime."
            assert isinstance(birth_date, (datetime, date)), txt
            if isinstance(birth_date, datetime):
                self._birth_date = birth_date.date()
            else:
                self._birth_date = birth_date
        else:
            self._birth_date = birth_date

    def getSurname(self):
        """get the participant surname"""
        return self._surname

    def getName(self):
        """get the participant name"""
        return self._name

    def getGender(self):
        """get the participant gender"""
        return self._gender

    def getHeight(self):
        """get the participant height in meter"""
        return self._height

    def getWeight(self):
        """get the participant weight in kg"""
        return self._weight

    def getBirthDate(self):
        """get the participant birth date"""
        return self._birth_date

    def getBMI(self):
        """get the participant BMI in kg/m^2"""
        return self._weight / (self._height**2)

    def getFullName(self):
        """
        get the participant full name.
        """
        return "{} {}".format(self._surname, self._name)

    def getAge(self, dt: date = None):
        """
        get the age of the participant in years

        Parameters
        ----------
        dt: datetime.date (optional)
            In case age is not directly available, this parameter is used
            to calulate the age of the participant starting from the birth date.

        Returns
        -------
        age: float
            the age of the participant.
        """
        if self._age is not None:
            return self._age
        if dt is None:
            dt = datetime.now().date()
        else:
            assert isinstance(dt, date), "'dt' must be a datetime.date object."
        return dt.year - self._birth_date.year

    def getMaxHR(self, dt: date = None):
        """
        get the maximum theoretical heart rate according to Gellish.

        Parameters
        ----------
        dt: datetime.date (optional)
            if provided. Age and maxHR are calculated considering the
            provided date and the stored birth date.

        Returns
        -------
        hr: float
            the calculated maximum heart rate.

        References
        ----------
        Gellish RL, Goslin BR, Olson RE, McDonald A, Russi GD, Moudgil VK.
            Longitudinal modeling of the relationship between age and maximal
            heart rate.
            Med Sci Sports Exerc. 2007;39(5):822-9.
            doi: 10.1097/mss.0b013e31803349c6.
        """
        if dt is None:
            dt = datetime.now().date()
        else:
            assert isinstance(dt, date), "'dt' must be a datetime.date object."
        return 207 - 0.7 * self.getAge(dt)

    @classmethod
    def fromCosmedOmnia(cls, df):
        """
        return the Participant object read by a Cosmed Omnia excel export.

        Parameters
        ----------
        df: pandas.DataFrame
            the dataframe resulting from the Cosmed Omnia exporting function.

        Returns
        -------
        p: Participant
            a Participant instance.
        """
        assert isinstance(df, (pd.DataFrame)), "'df' must be a pandas.DataFrame"
        surname, name, gender, _, height, weight, birth_date = df.iloc[:7, 1]
        birth_date = datetime.strptime(birth_date + "-00:00:00", datetime_format).date()
        return cls(surname, name, gender, height, weight, birth_date)

    def toDict(self, dt: date = None):
        """
        return a dict representation of self

        Parameters
        ----------
        dt: datetime.date (optional)
            if provided. Age and maxHR are calculated considering the
            provided date and the stored birth date.

        Returns
        -------
        out: dict
            a dict with all the data relative to the participant.
        """
        return {
            "FULL NAME": self.getFullName(),
            "SURNAME": self.getSurname(),
            "NAME": self.getName(),
            "GENDER": self.getGender(),
            "HEIGHT (m)": self.getHeight(),
            "WEIGHT (kg)": self.getWeight(),
            "BMI (kg/m^2)": self.getBMI(),
            "BIRTH DATE (dd/mm/aaaa)": self.getBirthDate(),
            "AGE (yrs)": self.getAge(),
            "MAX HR (bpm)": self.getMaxHR(),
        }

    def toDataFrame(self, dt: date = None):
        """
        return a pandas.DataFrame representation of self

        Parameters
        ----------
        dt: datetime.date (optional)
            if provided. Age and maxHR are calculated considering the
            provided date and the stored birth date.

        Returns
        -------
        out: pandas.DataFrame
            a dataframe with all the data relative to the participant.
        """
        return pd.DataFrame({i: [v] for i, v in self.toDict().items()})


#! FUNCTIONS


def magnitude(y, base=10):
    """
    return the order in the given base of the value

    Input:

        value (float)

            the value to be checked

        base (float)

            the base to be used to define the order of the number

    Output:

        mag (float)

            the number required to elevate the base to get the value
    """
    if y == 0 or base == 0:
        return 0
    else:
        return np.log(abs(y)) / np.log(base)


def get_files(path, extension="", check_subfolders=False):
    """
    list all the files having the required extension in the provided folder
    and its subfolders (if required).

    Parameters
    ----------
        path: str
            a directory where to look for the files.

        extension: str
            a str object defining the ending of the files that have to be
            listed.

        check_subfolders: bool
            if True, also the subfolders found in path are searched,
            otherwise only path is checked.

    Returns
    -------
        files: list
            a list containing the full_path to all the files corresponding
            to the input criteria.
    """

    # output storer
    out = []

    # surf the path by the os. walk function
    for root, dirs, files in os.walk(path):
        for obj in files:
            if obj[-len(extension) :] == extension:
                out += [os.path.join(root, obj)]

        # handle the subfolders
        if not check_subfolders:
            break

    # return the output
    return out


def to_excel(path, df, sheet="Sheet1", keep_index=True, new_file=False):
    """
    a shorthand function to save a pandas dataframe to an excel path

    Input:

        path (str)

            the path to the path where to store the path.

        data (pandas.DataFrame)

            a pandas.DataFrame.

        sheet (str)

            the sheet name.

        keep_index (boolean)

            if True, the dataframe index is preserved.
            Otherwise it is ignored.

        new_file (boolean)

            if True, a completely new path will be created.

    Output:

        The data stored to the indicated path.
    """

    # get the workbook
    if os.path.exists(path) and not new_file:
        wb = xl.load_workbook(path)
    else:
        wb = xl.Workbook()
        try:
            sh = wb["Sheet"]
            wb.remove(sh)
        except Exception:
            pass

    # get the sheet
    try:
        sh = wb[sheet]
        wb.remove(sh)
    except Exception:
        pass
    sh = wb.create_sheet(sheet)

    # write the headers
    [R, C] = df.shape
    if keep_index:
        index = np.atleast_2d(df.index.tolist())
        if index.shape[0] == 1:
            index = index.T
        data_cols = index.shape[1] + 1
    else:
        data_cols = 1
    header = np.atleast_2d(df.columns.tolist())
    if header.shape[0] == 1:
        header = header.T
    data_rows = header.shape[1] + 1
    for i, col in enumerate(header):
        for j, el in enumerate(col):
            ch = data_cols + i
            rh = 1 + j
            sh.cell(rh, ch, el)
    if keep_index:
        for i, row in enumerate(index):
            for j, el in enumerate(row):
                ri = data_rows + i
                ci = 1 + j
                sh.cell(ri, ci, el)

    # write the data
    V = df.values
    for r in range(R):
        for c in range(C):
            sh.cell(data_rows + r, data_cols + c, V[r, c])

    # save data
    wb.save(path)


def from_excel(path, sheets=None, **kwargs):
    """
    a shorthand function to collect data from an excel path
    and to store them into a dict.

    Input:

        path (str)

            the path to the path where to store the path.

        sheets (list of str)

            the name of the sheets to be imported. If None
            all sheets will be imported.

        kwargs

            additional arguments passed to pandas.read_excel

    Output:

        a dict object with keys equal to the sheets name and pandas
        dataframe as elements of each sheet in the excel path.
    """

    # retrive the data in the path path
    try:
        xlfile = pd.ExcelFile(path)
        sheets = xlfile.sheet_names if sheets is None else [sheets]
        sheets = np.array(sheets).flatten()
    except Exception:
        sheets = []
    finally:
        xlfile.close()

    # return the dict
    return {i: pd.read_excel(path, i, **kwargs) for i in sheets}


def get_time(tic=None, toc=None, as_string=True, compact=True):
    """
    get the days, hours, minutes and seconds between the two times.
    If only tic is provided, it is considered as the lapsed time.
    If neither tic nor toc are provided, the function returns the
    current time as float.

    Input (optional)

        tic (int)

            an integer representing the starting time

        toc (int)

            an integer indicating the stopping time

        as_string (bool)

            should the output be returned as string?

        compact (bool)

            if "as_string" is true, should the time be reported in a
            compact or in an extensive way?

    Output:

        If nothing is provided, the function returns the current time.
        If only tic is provided, the function returns the time value
        from it to now. If both tic and toc are provided, the function
        returns the time difference between them.
    """

    # check what to do
    if tic is None:
        return time.time()
    elif toc is None:
        tm = np.float(tic)
    else:
        tm = np.float(toc - tic)

    # convert the time value in days, hours, minutes,
    # seconds and milliseconds
    d = int(np.floor(tm / 86400))
    tm -= d * 86400
    h = int(np.floor(tm / 3600))
    tm -= h * 3600
    m = int(np.floor(tm / 60))
    tm -= m * 60
    s = int(np.floor(tm))
    tm -= s
    ms = int(np.round(1000 * tm, 0))

    # report the calculated time
    if not as_string:
        return {
            "Days": [d],
            "Hours": [h],
            "Minutes": [m],
            "Seconds": [s],
            "Milliseconds": [ms],
        }
    else:
        st = "{:0>2d}".format(d) + (" Days - " if not compact else ":")
        st += "{:0>2d}".format(h)
        st += " Hours - " if not compact else ":"
        st += "{:0>2d}".format(m)
        st += " Minutes - " if not compact else ":"
        st += "{:0>2d}".format(s)
        st += " Seconds - " if not compact else ":"
        st += "{:0>3d}".format(ms)
        st += " Milliseconds" if not compact else ""
        return st


def lvlup(path):
    """
    Goes to the superior level in the directory path.

    Input:

        path (str)

            a path or a directory. Otherwise a message is casted and
            the function returns the input as is.

    Output:

        s (str)

            a string reflecting the superior directory of path.
    """

    # return the upper level
    return os.path.sep.join(path.split(os.path.sep)[:-1])
