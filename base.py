# IMPORTS

import itertools as it
from typing import Union
import numpy as np
import openpyxl as xl
import os
import pandas as pd
import scipy.interpolate as si
import scipy.linalg as sl
import scipy.signal as ss
import sympy as sy
import time


# CLASSES


class LinearRegression:
    """
    Obtain the regression coefficients according to the Ordinary Least
    Squares approach.

    Parameters
    ----------
    y:  (samples, dimensions) numpy array or pandas.DataFrame
        the array containing the dependent variable.

    x:  (samples, features) numpy array or pandas.DataFrame
        the array containing the indipendent variables.

    fit_intercept: bool
        Should the intercept be included in the model?
        Otherwise it will be set to zero.
    """

    def __init__(
        self,
        y: Union[np.ndarray, pd.DataFrame],
        x: Union[np.ndarray, pd.DataFrame],
        fit_intercept: bool = True,
    ):

        # add the input parameters
        txt = "{} must be a {} object."
        assert isinstance(fit_intercept, bool), txt.format("fit_intercept", "bool")
        self.fit_intercept = fit_intercept

        # correct the shape of y and x
        self.Y = self._simplify(y)
        self.X = self._simplify(x)
        txt = "'X' and 'Y' number of rows must be identical."
        assert self.X.shape[0] == self.Y.shape[0], txt

        # add the ones for the intercept
        if self.fit_intercept:
            xx = self.X.values
            xx = np.hstack([np.ones((self.X.shape[0], 1)), xx])

        # get the coefficients and intercept
        self.betas = pd.DataFrame(
            data=sl.pinv(xx.T.dot(xx)).dot(xx.T).dot(self.Y.values),
            index=["INTERCEPT"] + self.X.columns.to_numpy().tolist(),
            columns=self.Y.columns.to_numpy().tolist(),
        )

    def _simplify(self, v):
        """
        internal method to check entries in the constructor.
        """
        txt = "the input object must be a pandas.DataFrame, a numpy.ndarray or None."
        if isinstance(v, pd.DataFrame):
            xx = v.astype(float)
        elif isinstance(v, np.ndarray):
            xx = pd.DataFrame(np.atleast_1d(v).astype(float))
        else:
            raise NotImplementedError(txt)
        return xx

    def predict(self, x):
        """
        predict the fitted Y value according to the provided x.
        """
        return self._predict(self._simplify(x))

    def _predict(self, xx):
        """
        predict the fitted Y value according to the provided x.
        """
        if self.fit_intercept:
            b0 = self.betas.iloc[0]
            bs = self.betas.iloc[1:]
            zz = xx.dot(bs) + b0
        else:
            zz = xx.dot(self.betas)
        return zz


class PolynomialRegression(LinearRegression):
    """
    Obtain the regression coefficients according to the Ordinary Least
    Squares approach.

    Parameters
    ----------
    y:  (samples, dimensions) numpy array or pandas.DataFrame
        the array containing the dependent variable.

    x:  (samples, 1) numpy array or pandas.DataFrame
        the array containing the indipendent variables.

    n: int
        the order of the polynome.

    fit_intercept: bool
        Should the intercept be included in the model?
        Otherwise it will be set to zero.
    """

    def __init__(
        self,
        y: Union[np.ndarray, pd.DataFrame],
        x: Union[np.ndarray, pd.DataFrame],
        n: int = 1,
        fit_intercept: bool = True,
    ):
        # add the input parameters
        txt = "{} must be a {} object."
        assert isinstance(fit_intercept, bool), txt.format("fit_intercept", "bool")
        self.fit_intercept = fit_intercept
        assert isinstance(n, int), txt.format("n", "int")
        self.n = n

        # correct the shape of y and x
        self.Y = self._simplify(y)
        self.X = self._simplify(x, self.n)
        txt = "'X' and 'Y' number of rows must be identical."
        assert self.X.shape[0] == self.Y.shape[0], txt

        # add the ones for the intercept
        if self.fit_intercept:
            xx = self.X.values
            xx = np.hstack([np.ones((self.X.shape[0], 1)), xx])

        # get the coefficients and intercept
        self.betas = pd.DataFrame(
            data=sl.pinv(xx.T.dot(xx)).dot(xx.T).dot(self.Y.values),
            index=["INTERCEPT"] + self.X.columns.to_numpy().tolist(),
            columns=self.Y.columns.to_numpy().tolist(),
        )

    def _simplify(self, v, n=None):
        """
        internal method to check entries in the constructor.
        """
        txt = "the input object must be a pandas.DataFrame, a numpy.ndarray or None."
        if isinstance(v, pd.DataFrame):
            xx = v.astype(float)
        elif isinstance(v, np.ndarray):
            xx = pd.DataFrame(np.atleast_1d(v).astype(float))
        else:
            raise NotImplementedError(txt)
        if n is not None:
            xx = pd.concat([xx ** (i + 1) for i in range(n)], axis=1)
            cols = ["C{}".format(i + 1) for i in range(xx.shape[1])]
            xx.columns = pd.Index(cols)
        return xx

    def predict(self, x):
        """
        predict the fitted Y value according to the provided x.
        """
        return self._predict(self._simplify(x, self.n))


class PowerRegression(LinearRegression):
    def __init__(self, y, x, digits=5):
        """
        Obtain the regression coefficients according to the power model:

                    y = b_0 * x_1 ^ b_1 * ... * x_n ^ b_n

        Input

            y (2D column numpy array)

                the array containing the dependent variable.

            x (2D column numpy array)

                the array containing the indipendent variable.
                The number of rows must equal the rows of y, while
                each column will be a regressor (i.e. an indipendent
                variable).

            digits (int)

                the number of digits used to print the coefficients
        """

        # correct the shape of y and x
        assert isinstance(digits, (int)), "'digits' must be and 'int'."
        self.digits = digits
        YY = np.log(self.__simplify__(y, "Y", None))
        XX = np.log(self.__simplify__(x, "X", None))
        txt = "'X' and 'Y' number of rows must be identical."
        assert XX.shape[0] == YY.shape[0], txt

        # add the ones for the intercept
        XX = np.hstack([np.ones((XX.shape[0], 1)), XX])

        # get the coefficients
        coefs = sl.pinv(XX.T.dot(XX)).dot(XX.T).dot(YY)
        coefs[0] = np.e ** coefs[0]

        # get the coefficients and intercept
        self._coefs = pd.DataFrame(
            data=coefs, index=self.__iv_labels__, columns=self.__dv_labels__
        )

        # obtain the symbolic representation of the equation
        self.symbolic = []
        for c, var in enumerate(self.betas):
            vrs = self.X.columns.to_numpy()
            a = self.betas[var].values[0]
            bs = self.betas[var].values[1:]
            line = sy.Float(a, digits)
            for v, b in zip(vrs, bs):
                line = line * sy.symbols(v) ** sy.Float(b, digits)
            self.symbolic += [sy.Eq(sy.symbols(var), line)]

    def copy(self):
        """
        copy the current instance
        """
        return PowerRegression(self.Y, self.X)

    def predict(self, x):
        """
        predict the fitted Y value according to the provided x.
        """
        X = self.__simplify__(x, "x", None)
        m = self.betas.shape[0] - 1
        assert X.shape[1] == m, "'X' must have {} columns.".format(m)
        Z = []
        for dim in self.betas:
            coefs = self.betas[dim].values.T
            Z += [np.prod(X ** coefs[1:], axis=1) * coefs[0]]
        Z = pd.DataFrame(np.atleast_2d(Z).T)
        if isinstance(x, pd.DataFrame):
            idx = x.index
        else:
            idx = pd.Index(np.arange(X.shape[0]))
        return pd.DataFrame(Z, index=idx, columns=self.__dv_labels__)

    @property
    def __iv_labels__(self):
        """
        return the labels for the regressors.
        """
        lbls = ["b{}".format(i + 1) for i in np.arange(len(self.X.columns))]
        return ["b0"] + lbls


class HyperbolicRegression(LinearRegression):
    def __init__(self, y, x, digits=5):
        """
        Obtain the regression coefficients according to the (Rectangular) Least
        Squares Hyperbolic function:
                                                            b * x
                        (x + a) * (y + b) = a * b ==> y = - -----
                                                            a + x
        Input

            y (2D column numpy array)

                the array containing the dependent variable.

            x (2D column numpy array)

                the array containing the indipendent variable.
                The number of rows must equal the rows of y, while
                each column will be a regressor (i.e. an indipendent
                variable).

            digits (int)

                the number of digits used to render the coefficients.
        """

        # add the input parameters
        txt = "{} must be a {} object."
        assert isinstance(digits, (int)), txt.format("digits", "int")
        self.digits = digits

        # correct the shape of y and x and get their reciprocal values
        YY = self.__simplify__(y, "Y", None) ** (-1)
        XX = self.__simplify__(x, "X", None) ** (-1)
        txt = "'X' and 'Y' number of rows must be identical."
        assert XX.shape[0] == YY.shape[0], txt
        assert XX.shape[1] == 1, "'X' must have just 1 column."

        # add the ones for the intercept
        XX = np.hstack([np.ones((XX.shape[0], 1)), XX])

        # get the linear regression coefficients
        coefs = sl.pinv(XX.T.dot(XX)).dot(XX.T).dot(YY)

        # obtain the hyberbolic coefficients
        # a = -1 / intercept
        # b =  slope / intercept
        coefs = [[coefs[1][0] / coefs[0][0]], [-1 / coefs[0][0]]]
        self._coefs = pd.DataFrame(
            data=coefs, index=self.__iv_labels__, columns=self.__dv_labels__
        )

        # obtain the symbolic representation of the equation
        x = sy.symbols(self.X.columns.to_numpy()[0])
        c = [sy.Float(i, self.digits) for i in self.betas.values]
        y = sy.symbols(self.betas.columns.to_numpy()[0])
        self.symbolic = [sy.Eq(y, (c[1] * x) / (c[0] + x))]

    def copy(self):
        """
        copy the current instance
        """
        return HyperbolicRegression(self.Y, self.X)

    def predict(self, x):
        """
        predict the fitted Y value according to the provided x.
        """
        X = self.__simplify__(x, "x", None)
        assert X.shape[1] == 1, "'X' must have 1 column."
        Z = -self.betas.loc["b"].values * X / (self.betas.loc["a"].values + X)
        if isinstance(x, pd.DataFrame):
            idx = x.index
        else:
            idx = pd.Index(np.arange(X.shape[0]))
        return pd.DataFrame(Z, index=idx, columns=self.__dv_labels__)

    @property
    def __iv_labels__(self):
        """
        return the labels for the regressors.
        """
        return ["a", "b"]

    def to_string(self, digits=3):
        """
        Create a textual representation of the fitted model.

        Input:
            digits: (int)
                    the number of digits to be used to represent the
                    coefficients.

        Output:
            txt:    (str)
                    a single string representing the fitted model.
        """
        frm = "{:+." + str(digits) + "f}"
        txt = "({} " + frm + ") * ({} " + frm + ") = " + frm
        return txt.format(
            self.betas.columns.to_numpy()[0],
            self.betas.loc["a"].values.flatten()[0],
            self.__dv_labels__()[0],
            self.betas.loc["b"].values.flatten()[0],
            np.prod(self.betas.values.flatten()),
        )


# FUNCTIONS


def d1y(y, x=None, dt=1):
    """
    return the first derivative of y.

    Parameters
    ----------

    y: ndarray with shape [n,]
        the signal to be derivated

    x: ndarray with shape [n,] or None
        the optional signal from which y has to  be derivated

    dt: float
        the difference between samples in y.
        NOTE: if x is provided, this parameter is ignored

    Returns
    -------

    z: ndarray
        an array being the second derivative of y

    References
    ----------

    Winter DA. Biomechanics and Motor Control of Human Movement. Fourth Ed.
        Hoboken, New Jersey: John Wiley & Sons Inc; 2009.
    """

    # get x
    if x is None:
        x = np.arange(len(y)) * dt

    # get the derivative
    return (y[2:] - y[:-2]) / (x[2:] - x[:-2])


def d2y(y, x=None, dt=1):
    """
    return the second derivative of y.

    Parameters
    ----------

    y: ndarray with shape [n,]
        the signal to be derivated

    x: ndarray with shape [n,] or None
        the optional signal from which y has to  be derivated

    dt: float
        the difference between samples in y.
        NOTE: if x is provided, this parameter is ignored

    Returns
    -------

    z: ndarray
        an array being the second derivative of y

    References
    ----------

    Winter DA. Biomechanics and Motor Control of Human Movement. Fourth Ed.
        Hoboken, New Jersey: John Wiley & Sons Inc; 2009.
    """

    # data check
    txt = "{} must be an object of class {}."
    assert isinstance(y, np.ndarray), txt.format("y", "ndarray")
    assert y.ndim == 1, "y must be a 1D array."
    assert isinstance(dt, (int, float)), txt.format("dt", "(int, float)")
    if x is None:
        x = np.arange(len(y)) * dt
    assert isinstance(x, np.ndarray), txt.format("x", "ndarray")
    assert x.ndim == 1, "x must be a 1D array."

    # get the derivative
    dy = (y[2:] - y[1:-1]) / (x[2:] - x[1:-1])
    dy -= (y[1:-1] - y[:-2]) / (x[1:-1] - x[:-2])
    dx = (x[2:] - x[:-2]) * 0.5
    return dy / dx


def freedman_diaconis_bins(y):
    """
    return a digitized version of y where each value is linked to a
    bin (i.e an int value) according to the rule.

                             IQR(x)
            width = 2 * ---------------
                        len(x) ** (1/3)

    Input:

        y (1D array)

            a ndarray that has to be digitized.

    Output:

        d (1D array)

            an array with the same shape of y containing the index
            of the bin of which the corresponding sample of y is part.

    References:

        Freedman D, Diaconis P.
            (1981) On the histogram as a density estimator:L 2 theory.
            Z. Wahrscheinlichkeitstheorie verw Gebiete 57: 453–476.
            doi: 10.1007/BF01025868
    """

    # y IQR
    q1 = np.quantile(y, 0.25)
    q3 = np.quantile(y, 0.75)
    iqr = q3 - q1

    # get the width
    w = 2 * iqr / (len(y) ** (1 / 3))

    # get the number of intervals
    n = int(np.floor(1 / w)) + 1

    # digitize z
    d = np.zeros(y.shape)
    for i in np.arange(n) + 1:
        d[np.argwhere((y >= (i - 1) * w) & (y < i * w)).flatten()] = i - 1
    return d


def mean_filter(signal, n=1, pad_style="reflect", offset=0.5):
    """
    apply a moving average filter to the signal.

    Parameters
    ----------

    signal: 1D array
        the signal to be filtered.

    n: int
        the number of samples to be considered as averaging window.

    pad_style: "reflect"
        padding style mode. The the numpy.pad function is used. Please refer to the corresponding documentation
        for a detailed description of the padding modes.

    offset: float
        a value within the [0, 1] range defining how the averaging window is obtained. Offset=0, indicate that
        for each sample, the filtered value will be the mean of the subsequent n-1 values plus the current sample.
        Offset=1, on the other hand, calculates the filtered value at each sample as the mean of the n-1 preceding
        values plus the current sample.
        Offset=0.5, centers the averaging window around the actual sample being evaluated.

    Returns
    -------

    z: 1D array
        The filtered signal.
    """

    # data check
    txt = "{} must be an object of class {}."
    assert isinstance(signal, np.ndarray), txt.format("signal", "ndarray")
    assert signal.ndim == 1, "signal must be a 1D array."
    assert isinstance(n, int), txt.format("n", "int")
    assert isinstance(pad_style, str), txt.format("pad_style", "str")
    assert isinstance(offset, (int, float)), txt.format("offset", "(int, float)")
    assert 0 <= offset <= 1, "offset must be a float in the [0, 1] range."

    # get the window range
    w = np.unique((np.arange(n) - offset * (n - 1)).astype(int))

    # get the indices of the samples
    i = [w + n - 1 + j for j in np.arange(len(signal))]

    # padding
    y = np.pad(signal, n - 1, pad_style)

    # get the mean of each window
    return np.array([np.mean(y[j]) for j in i]).flatten()


def median_filter(signal, n=1, pad_style="reflect", offset=0.5):
    """
    apply a median filter to the signal.

    Parameters
    ----------

    signal: 1D array
        the signal to be filtered.

    n: int
        the number of samples to be considered as averaging window.

    pad_style: "reflect"
        padding style mode. The the numpy.pad function is used. Please refer to the corresponding documentation
        for a detailed description of the padding modes.

    offset: float
        a value within the [0, 1] range defining how the window is obtained. Offset=0, indicate that
        for each sample, the filtered value will be the median of the subsequent n-1 values plus the current sample.
        Offset=1, on the other hand, calculates the filtered value at each sample as the median of the n-1 preceding
        values plus the current sample.
        Offset=0.5, centers the window around the actual sample being evaluated.

    Returns
    -------

    z: 1D array
        The filtered signal.
    """

    # data check
    txt = "{} must be an object of class {}."
    assert isinstance(signal, np.ndarray), txt.format("signal", "ndarray")
    assert signal.ndim == 1, "signal must be a 1D array."
    assert isinstance(n, int), txt.format("n", "int")
    assert isinstance(pad_style, str), txt.format("pad_style", "str")
    assert isinstance(offset, (int, float)), txt.format("offset", "(int, float)")
    assert (
        0 <= offset <= 1
    ), "offset must be a numeric values included in the [0, 1] range."

    # get the window range
    w = np.unique((np.arange(n) - offset * (n - 1)).astype(int))

    # get the indices of the samples
    i = [w + n - 1 + j for j in np.arange(len(signal))]

    # padding
    y = np.pad(signal, n - 1, pad_style)

    # get the mean of each window
    return np.array([np.median(y[j]) for j in i]).flatten()


def interpolate_cs(y, n=None, x_old=None, x_new=None):
    """
    Get the cubic spline interpolation of y.

    Parameters
    ----------

    y: 1D array
        the data to be interpolated.

    n: int, None
        the number of points for the interpolation.

    x_old: 1D array, None
        the x coordinates corresponding to y. It is ignored if n is provided.

    x_new: 1D array, None
        the newly (interpolated) x coordinates corresponding to y.
        It is ignored if n is provided.

    Returns
    -------

    z: 1D array
        the interpolated y axis
    """

    # control of the inputs
    txt = "{} must be an object of class {}."
    assert isinstance(y, np.ndarray), txt.format("y", "ndarray")
    assert y.ndim == 1, "y must be a 1D array."
    if n is not None:
        assert isinstance(n, int), txt.format("n", "int")
        x_old = np.arange(len(y))
        x_new = np.linspace(np.min(x_old), np.max(x_old), n)
    else:
        assert isinstance(x_old, np.ndarray), txt.format("x_old", "ndarray")
        assert x_old.ndim == 1, "x_old must be a 1D array."
        assert len(y) == len(x_old), "x_old and y must have the same len."
        assert isinstance(x_new, np.ndarray), txt.format("x_new", "ndarray")
        assert x_new.ndim == 1, "x_new must be a 1D array."

    # get the cubic-spline interpolated y
    cs = si.CubicSpline(x_old, y)
    return cs(x_new)


def residuals_analysis(
    signal,
    fs,
    f_num=1000,
    f_max=None,
    segments=2,
    min_samples=2,
    which_segment=None,
    filt_fun=None,
    filt_opt=None,
):
    """
    Perform Winter's residual analysis of y.

    Parameters
    ----------

    signal: 1D array
        the signal to be investigated

    fs: float
        the sampling frequency of the signal.

    f_num: int, optional
        the number of frequencies to be tested within the (0, f_max) range to create the residuals curve of
        the Winter's residuals analysis approach.

    f_max: float, optional
        the maximum filter frequency that is tested. If None, it is defined as the frequency covering the 99%
        of the cumulative signal power.

    segments: int, optional
        the number of segments that can be used to fit the residuals curve in order to identify the best
        deflection point.
        NOTE: values above 3 will greatly increase the computation time.

    min_samples: int, optional
        the minimum number of elements that have to be considered for each segment during the calculation of
        the best deflection point.

    which_segment: int, optional
        the segment to be considered as the one allowing the calculation of the optimal cut-off.
        It must be an int in the [1, segments] range. If None, the segment resulting in the most flat line from
        those that have been calculated is used.

    filt_fun: function(signal, frequency, **kwargs)
        the filter to be used for the analysis. If None, a Butterworth, low-pass, 4th order phase-corrected
        filter is used. If a function is provided, two arguments are mandatory:

        - signal, 1D array passed as first argument
        - frequency, a float (positive) number passed as second argument.
        - additional keyworded parameters that are directly passed to the function.

        This function must return an array with the same shape of signal being its filtered copy with cut-off
        frequency equal to "frequency".

    filt_opt: dict, optional
        the options for the filter. If not None, a dict containing the key-values combinations to
        be passed to filt_fun.

    Returns
    -------

    cutoff: float
        the suggested cutoff value

    SSEs: pandas.DataFrame
        a pandas.DataFrame with the selected frequencies as index and the Sum of Squared Residuals as columns.

    Notes
    -----

    The signal is filtered over a range of frequencies and the sum of squared residuals (SSE) against the original
    signal is computer for each tested cut-off frequency. Next, a series of fitting lines are used to estimate the
    optimal disruption point defining the cut-off frequency optimally discriminating between noise and good quality
    signal.

    References
    ----------

    Winter DA 2009, Biomechanics and Motor Control of Human Movement. Fourth Ed.
        John Wiley & Sons Inc, Hoboken, New Jersey (US).

    Lerman PM 1980, Fitting Segmented Regression Models by Grid Search. Appl Stat. 29(1):77.
    """

    # data check
    txt = "{} must be an object of class {}."
    assert isinstance(signal, np.ndarray), txt.format("signal", "ndarray")
    assert signal.ndim == 1, "signal must be a 1D array."
    assert isinstance(fs, (int, float)), txt.format("fs", "(int, float)")
    assert isinstance(f_num, int), txt.format("f_num", "int")
    assert f_num > 1, "'f_num' must be > 1."
    if f_max is None:
        P, F = psd(signal, fs)
        f_max = np.argwhere(np.cumsum(P) / np.sum(P) >= 0.99).flatten()
        f_max = np.min([fs / 2, F[f_max[0]]])
    assert isinstance(f_max, (int, float)), txt.format("f_max", "(int, float, None)")
    assert isinstance(min_samples, int), txt.format("min_samples", "int")
    assert min_samples >= 2, "'min_samples' must be >= 2."
    if which_segment is not None:
        seg_txt = "'which_segment' must be an int in the [1, {}] range."
        assert 1 <= which_segment < segments, seg_txt.format(segments)
    if filt_fun is None:
        filt_fun = butt_filt
    if filt_opt is None:
        filt_opt = {"n": 4, "fs": fs, "type": "lowpass", "phase_corrected": True}

    # get the frequency span
    freqs = np.linspace(0, f_max, f_num + 1)[1:]

    # get the SSEs
    Q = np.array(
        [np.sum((signal - filt_fun(signal, i, **filt_opt)) ** 2) for i in freqs]
    )

    # reshape the SSE as dataframe
    D = pd.DataFrame(Q, index=freqs, columns=["SSE"])

    # get the optimal crossing over point that separates the S regression
    # lines best fitting the residuals data.
    F = crossovers(Q, segments, min_samples)[1]

    # get the intercept of optimal line
    # if which_segment is None, find the fitting line being the most flat
    if which_segment is None:
        I = F[np.argmin([abs(i[0]) for i in F])][1]
    else:
        I = F[which_segment][1]

    # get the optimal cutoff
    opt = freqs[np.argmin(abs(Q - I))]

    # get the fitting lines for each segment
    for i, fit in enumerate(F):
        D.insert(D.shape[1], "S{}".format(i + 1), fit[1] * freqs + fit[0])

    # return the parameters
    return opt, D


def crossovers(signal, segments=2, min_samples=5):
    """
    Detect the position of the crossing over points between K regression lines used to best fit the data.

    Parameters
    ----------

    signal: 1D array
        the signal to be fitted.

    segments: int, optional
        the number of segments that can be used to fit the residuals curve in order to identify the best
        deflection point.
        NOTE: values above 3 will greatly increase the computation time.

    min_samples: int, optional
        the minimum number of elements that have to be considered for each segment during the calculation of the
        best deflection point.

    Returns
    -------

    crossings: 1D array
        An ordered array of indices containing the samples corresponding to the detected crossing over points.

    coefs: list
        A list of tuples containing the slope and intercept of the line describing each fitting segment.

    Notes
    -----

    the steps involved in the calculations can be summarized as follows:

        1)  Get all the segments combinations made possible by the given number of crossover points.

        2)  For each combination, calculate the regression lines corresponding to each segment.

        3)  For each segment calculate the residuals between the calculated regression line and the effective data.

        5)  Once the sum of the residuals have been calculated for each combination, sort them by residuals amplitude.

    References
    ----------

    Lerman PM 1980, Fitting Segmented Regression Models by Grid Search. Appl Stat. 29(1):77.
    """

    # control the inputs
    txt = "{} must be an object of class {}."
    assert isinstance(signal, np.ndarray), txt.format("signal", "ndarray")
    assert signal.ndim == 1, "signal must be a 1D array."
    assert isinstance(segments, int), txt.format("segments", "int")
    assert isinstance(min_samples, int), txt.format("min_samples", "int")
    assert min_samples >= 2, "'min_samples' must be >= 2."

    # get the residuals calculating formula
    def SSEs(x, y, s):
        """
        method used to calculate the residuals

        Parameters
        ----------

        x: 1D array
            the x axis signal

        y: 1D array
            the y axis signal

        s: list
            the extremes among which the segments have to be fitted

        Returns
        -------

        sse: float
            the sum of squared errors corresponding to the error obtained fitting the y-x relationship according to
            the segments provided by s.
        """
        c = [
            np.arange(s[i], s[i + 1] + 1) for i in np.arange(len(s) - 1)
        ]  # get the coordinates
        z = [
            np.polyfit(x[i], y[i], 1) for i in c
        ]  # get the fitting parameters for each interval
        v = [
            np.polyval(v, x[c[i]]) for i, v in enumerate(z)
        ]  # get the regression lines for each interval
        return np.sum(
            [np.sum((y[c[i]] - v) ** 2) for i, v in enumerate(v)]
        )  # get the sum of squared residuals

    # get the X axis
    x = np.arange(len(signal))

    # get all the possible combinations of segments
    J = [
        np.arange(min_samples * i, len(signal) - min_samples * (segments - i))
        for i in np.arange(1, segments)
    ]
    J = [j for j in it.product(*J)]

    # remove those combinations having segments shorter than "samples"
    J = [i for i in J if np.all(np.diff(i) >= min_samples)]

    # generate the crossovers matrix
    J = np.hstack(
        (
            np.zeros((len(J), 1)),
            np.atleast_2d(J),
            np.ones((len(J), 1)) * len(signal) - 1,
        )
    ).astype(int)

    # calculate the residuals for each combination
    R = np.array([SSEs(x, signal, i) for i in J])

    # sort the residuals
    T = np.argsort(R)

    # get the optimal crossovers order
    O = x[J[T[0]]]

    # get the fitting slopes
    F = [np.arange(i0, i1) for i0, i1 in zip(O[:-1], O[1:])]
    F = [np.polyfit(i, signal[i], 1) for i in F]

    # return the crossovers
    return O[1:-1], F


def butt_filt(signal, fc, fs, n=4, type="lowpass", phase_corrected=True):
    """
    Provides a convenient function to call a Butterworth filter with the specified parameters.

    Parameters
    ----------

    signal: 1D array
        the signal to be filtered.

    fc: int, float, or 2 elements iterable of int/float values
        the filter cutoff in Hz.

    fs: int, float
        the sampling frequency of the signal in Hz.

    n: int, optional
        the order of the filter

    type: str, optional
        a string defining the type of the filter ("lowpass", "highpass", "bandpass", "stopband").

    phase_corrected: bool, optional
        should the filter be applied twice in opposite directions to correct for phase lag?

    Returns
    -------

    z: 1D array
        the resulting 1D filtered signal.
    """

    # control the inputs
    txt = "{} must be an object of class {}."
    assert isinstance(signal, (pd.Series, np.ndarray)), txt.format(
        "signal", "(pandas.Series, numpy.ndarray)"
    )
    assert signal.ndim == 1, "signal must be a 1D array."
    assert isinstance(fs, (int, float)), txt.format("fs", "(int, float)")
    if isinstance(fc, (np.ndarray, list)):
        txt2 = "'cutoff' length must be a list or numpy array of length 2."
        assert len(fc) == 2, txt2
        txt3 = "all elements in fc must be int or float."
        assert np.all([isinstance(i, (int, float)) for i in fc]), txt3
    else:
        assert isinstance(fc, (int, float)), txt.format("fc", "(int, float)")
    assert isinstance(type, str), txt.format("type", "str")
    types = ["lowpass", "highpass", "bandpass", "stopband"]
    assert np.any([type == i for i in types]), "type must be any of " + str(types)
    assert isinstance(phase_corrected, bool), txt.format("phase_corrected", "bool")

    # get the filter coefficients
    sos = ss.butter(n, (np.array([fc]).flatten() / (0.5 * fs)), type, output="sos")

    # get the filtered data
    return ss.sosfiltfilt(sos, signal) if phase_corrected else ss.sosfilt(sos, signal)


def psd(signal, fs=1):
    """
    compute the power spectrum of y using fft

    Parameters
    ----------

    signal: 1D array
        A 1D numpy array

    fs: int, float, optional
        the sampling frequency (in Hz) of the signal. If not provided the power spectrum frequencies are provided
        as normalized values within the (0, 0.5) range.

    Returns
    -------

    p: 1D array
        the power of each frequency

    k: 1D array
        the frequency corresponding to each element of pow.
    """

    # check the input
    txt = "{} must be an object of class {}."
    assert isinstance(signal, np.ndarray), txt.format("signal", "numpy.ndarray")
    assert signal.ndim == 1, "signal must be a 1D array."
    assert isinstance(fs, (int, float)), txt.format("fs", "(int, float)")

    # get the psd
    f = np.fft.rfft(signal - np.mean(signal)) / len(
        signal
    )  # normalized frequency spectrum
    a = abs(f)  # amplitude
    p = np.concatenate([[a[0]], 2 * a[1:-1], [a[-1]]]).flatten() ** 2  # power spectrum
    k = np.linspace(0, fs / 2, len(p))  # frequencies

    # return the data
    return p, k


def find_peaks(y, height=None):
    """
    detect the location (in sample units) of the peaks within y.

    Parameters
    ----------

    y: 1D array
        a 1D signal

    height: float, None
        a scalar defining the minimum height to be considered for a valid peak.
        If None, all peaks are returned.

    Returns
    -------

    p: 1D array
        the indices of the peaks (in sample units) in y.
    """

    # control the inputs
    txt = "{} must be an object of class {}."
    assert isinstance(y, np.ndarray), txt.format("y", "ndarray")
    assert y.ndim == 1, "y must be a 1D array."
    if height is None:
        height = np.min(y)
    assert isinstance(height, (int, float)), txt.format("height", "(int, float)")

    # get the first derivative of the signal
    d1 = d1y(y)

    # get the sign of the first derivative
    d1[d1 == 0] = 1
    sn = d1 / abs(d1)

    # get the peaks
    zc = np.argwhere(sn[1:] - sn[:-1] == -2).flatten()

    # exclude all peaks below the set height
    return zc[y[zc] >= height] + 1


def crossings(y, value=0.0):
    """
    Dectect the crossing points in x compared to value.

    Parameters
    ----------

        y: 1D array
            the 1D signal from which the crossings have to be found.

        value: float or 1D array with the same shape of y
            the value/s to be used to detect the crossings.
            If value is an array, the function will find those points in x
            crossing values according to the value of values at a
            given location.

    Returns
    -------

        c: 1D array
            the samples corresponding to the crossings.

        s: 1D array
            the sign of the crossings. Positive sign means crossings
            where the signal moves from values lower than "value" to
            values higher than "value". Negative sign indicate the
            opposite trend.
    """

    # get the sign of the signal without the offset
    sn = y - value
    sn[sn == 0.0] = 1
    sn = sn / abs(sn)

    # get the location of the crossings
    cr = np.argwhere(abs(sn[1:] - sn[:-1]) == 2).flatten()

    # return the crossings
    return cr, sn[cr]


def xcorr(x, y=None, biased=False, full=False):
    """
    set the (multiple) auto/cross correlation of the data in y.

    Parameters
    ----------
    x   1D array
        the signal from which the auto or cross-correlation is provided.

    y   1D array or None
        the signal from which the auto or cross-correlation is provided.
        if None. The autocorrelation of x is provided. Otherwise the x-y
        cross-correlation is returned.

    biased  bool
        if True, the biased auto/cross-correlation is provided.
        Otherwise, the 'unbiased' estimator is returned.

    full (bool)
        Should the negative lags be reported?

    Returns
    -------
    xcr 1D array
        the auto/cross-correlation value.

    lag 1D array
        the lags in sample units.
    """

    # take the autocorrelation if only y is provided
    if y is None:
        X = np.atleast_2d(x)
        Z = np.vstack([X, X])

    # take the cross-correlation (ensure the shortest signal is zero-padded)
    else:
        X = np.zeros((1, max(len(x), len(y))))
        Y = np.zeros((1, max(len(x), len(y))))
        X[:, : len(x)] = x
        Y[:, : len(y)] = y
        Z = np.vstack([X, Y])

    # get the matrix shape
    P, N = Z.shape

    # remove the mean from each dimension
    V = Z - np.atleast_2d(np.mean(Z, 1)).T

    # take the cross-correlation
    xc = []
    for i in np.arange(P - 1):
        for j in np.arange(i + 1, P):
            R = np.atleast_2d(ss.fftconvolve(V[i], V[j][::-1], "full"))
            xc = np.vstack((xc, R)) if len(xc) > 0 else np.copy(R)

    # average over all the multiples
    xc = np.mean(xc, 0)

    # adjust the output
    lags = np.arange(-(N - 1), N)
    if not full:
        xc = xc[(N - 1) :]
        lags = lags[(N - 1) :]

    # normalize
    xc /= (N + 1 - abs(lags)) if not biased else (N + 1)

    # return the cross-correlation data
    return xc, lags


def magnitude(y, base=10):
    """
    return the order in the given base of the value

    Input:

        value (float)

            the value to be checked

        base (float)

            the base to be used to define the order of the number

    Output:

        mag (float)

            the number required to elevate the base to get the value
    """
    if y == 0 or base == 0:
        return 0
    else:
        return np.log(abs(y)) / np.log(base)


def get_files(path, extension="", check_subfolders=False):
    """
    list all the files having the required extension in the provided folder and its subfolders (if required).

    Parameters
    ----------
        path: str
            a directory where to look for the files.

        extension: str
            a str object defining the ending of the files that have to be listed.

        check_subfolders: bool
            if True, also the subfolders found in path are searched, otherwise only path is checked.

    Returns
    -------
        files: list
            a list containing the full_path to all the files corresponding to the input criteria.
    """

    # output storer
    out = []

    # surf the path by the os. walk function
    for root, dirs, files in os.walk(path):
        for obj in files:
            if obj[-len(extension) :] == extension:
                out += [os.path.join(root, obj)]

        # handle the subfolders
        if not check_subfolders:
            break

    # return the output
    return out


def to_excel(path, df, sheet="Sheet1", keep_index=True, new_file=False):
    """
    a shorthand function to save a pandas dataframe to an excel path

    Input:

        path (str)

            the path to the path where to store the path.

        data (pandas.DataFrame)

            a pandas.DataFrame.

        sheet (str)

            the sheet name.

        keep_index (boolean)

            if True, the dataframe index is preserved.
            Otherwise it is ignored.

        new_file (boolean)

            if True, a completely new path will be created.

    Output:

        The data stored to the indicated path.
    """

    # get the workbook
    if os.path.exists(path) and not new_file:
        wb = xl.load_workbook(path)
    else:
        wb = xl.Workbook()
        try:
            sh = wb["Sheet"]
            wb.remove(sh)
        except Exception:
            pass

    # get the sheet
    try:
        sh = wb[sheet]
        wb.remove(sh)
    except Exception:
        pass
    sh = wb.create_sheet(sheet)

    # write the headers
    [R, C] = df.shape
    if keep_index:
        index = np.atleast_2d(df.index.tolist())
        if index.shape[0] == 1:
            index = index.T
        data_cols = index.shape[1] + 1
    else:
        data_cols = 1
    header = np.atleast_2d(df.columns.tolist())
    if header.shape[0] == 1:
        header = header.T
    data_rows = header.shape[1] + 1
    for i, col in enumerate(header):
        for j, el in enumerate(col):
            ch = data_cols + i
            rh = 1 + j
            sh.cell(rh, ch, el)
    if keep_index:
        for i, row in enumerate(index):
            for j, el in enumerate(row):
                ri = data_rows + i
                ci = 1 + j
                sh.cell(ri, ci, el)

    # write the data
    V = df.values
    for r in range(R):
        for c in range(C):
            sh.cell(data_rows + r, data_cols + c, V[r, c])

    # save data
    wb.save(path)


def from_excel(path, sheets=None, **kwargs):
    """
    a shorthand function to collect data from an excel path
    and to store them into a dict.

    Input:

        path (str)

            the path to the path where to store the path.

        sheets (list of str)

            the name of the sheets to be imported. If None
            all sheets will be imported.

        kwargs

            additional arguments passed to pandas.read_excel

    Output:

        a dict object with keys equal to the sheets name and pandas
        dataframe as elements of each sheet in the excel path.
    """

    # retrive the data in the path path
    try:
        xlfile = pd.ExcelFile(path)
        sheets = np.array(xlfile.sheet_names if sheets is None else [sheets]).flatten()
    except Exception:
        sheets = []
    finally:
        xlfile.close()

    # return the dict
    return {i: pd.read_excel(path, i, **kwargs) for i in sheets}


def get_time(tic=None, toc=None, as_string=True, compact=True):
    """
    get the days, hours, minutes and seconds between the two times.
    If only tic is provided, it is considered as the lapsed time.
    If neither tic nor toc are provided, the function returns the
    current time as float.

    Input (optional)

        tic (int)

            an integer representing the starting time

        toc (int)

            an integer indicating the stopping time

        as_string (bool)

            should the output be returned as string?

        compact (bool)

            if "as_string" is true, should the time be reported in a
            compact or in an extensive way?

    Output:

        If nothing is provided, the function returns the current time.
        If only tic is provided, the function returns the time value
        from it to now. If both tic and toc are provided, the function
        returns the time difference between them.
    """

    # check what to do
    if tic is None:
        return time.time()
    elif toc is None:
        tm = np.float(tic)
    else:
        tm = np.float(toc - tic)

    # convert the time value in days, hours, minutes,
    # seconds and milliseconds
    d = int(np.floor(tm / 86400))
    tm -= d * 86400
    h = int(np.floor(tm / 3600))
    tm -= h * 3600
    m = int(np.floor(tm / 60))
    tm -= m * 60
    s = int(np.floor(tm))
    tm -= s
    ms = int(np.round(1000 * tm, 0))

    # report the calculated time
    if not as_string:
        return {
            "Days": [d],
            "Hours": [h],
            "Minutes": [m],
            "Seconds": [s],
            "Milliseconds": [ms],
        }
    else:
        st = "{:0>2d}".format(d) + (" Days - " if not compact else ":")
        st += "{:0>2d}".format(h)
        st += " Hours - " if not compact else ":"
        st += "{:0>2d}".format(m)
        st += " Minutes - " if not compact else ":"
        st += "{:0>2d}".format(s)
        st += " Seconds - " if not compact else ":"
        st += "{:0>3d}".format(ms)
        st += " Milliseconds" if not compact else ""
        return st


def lvlup(path):
    """
    Goes to the superior level in the directory path.

    Input:

        path (str)

            a path or a directory. Otherwise a message is casted and
            the function returns the input as is.

    Output:

        s (str)

            a string reflecting the superior directory of path.
    """

    # return the upper level
    return os.path.sep.join(path.split(os.path.sep)[:-1])
