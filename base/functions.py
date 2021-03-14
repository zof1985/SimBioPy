
# IMPORTS

import os
import re
import traceback
import time
import numpy as np
import pandas as pd
import itertools as it
import scipy.signal as ss
import scipy.interpolate as si
import openpyxl as xl
from sklearn.model_selection import GridSearchCV
from sklearn.svm import SVR



# FUNCTIONS

def _validate_arr(y, dim = 1, shape = None, type = (int, float)):
    """
    private method used to validate numpy arrays.

    Input:

        y (Object)

            the object to be validated.

        dim (None or int)

            the number of dimensions that y must have.

        shape (None or iterable of int)

            the size that y must satisfy on each dimension.

        type (None or instance)

            the instance of the object.
    """

    # get the variable name
    stack = traceback.extract_stack()
    filename, lineno, function_name, code = stack[-2]
    name = re.compile(r'\((.*?)\).*$').search(code).groups()[0]

    # check y is an array
    txt = "'{}' must be a numpy array.".ormat(name)
    assert isinstance(y, (np.ndarray)), txt

    if type is not None:
        type = np.array([type]).flatten()
        assert y.dtype in type, "'type' must be any of {}".format(type)

    if shape is not None:
        txt = "'shape' must be an iterable of int values."
        assert np.all([isinstance(i, (int)) for i in shape]), txt

        for i in range(len(shape)):
            txt2 = "'{}.shape[{}]' must be {}.".format(name, i, shape[i])
            assert y.shape[i] == shape[i], txt2

    if dim is not None:
        assert isinstance(dim, (int)), "'dim' must be an int."
        txt = "'{}' must be a {} dimensional array.".format(name, dim)
        assert y.ndim == dim, txt



def _validate_obj(y, type):
    """
    private method used to validate numpy arrays.

    Input:

        y (Object)

            the object to be validated.

        type (iterable of obj instances)

            the instance(s) allowed for the object.
    """

    # get the variable name
    stack = traceback.extract_stack()
    filename, lineno, function_name, code = stack[-2]
    name = re.compile(r'\((.*?)\).*$').search(code).groups()[0]

    # check its type
    if type is not None:
        type = np.array([type]).flatten()
        assert y.dtype in type, "'{}' must be any of {}".format(name, type)



def d1y(y, x = None, dt = 1):
    '''
    return the first derivative of y.

    Input:

        y (ndarray with shape [n,])

            the signal to be derivated

        x (ndarray with shape [n,] or None)

            the optional signal from which y has to  be derivated

        dt (float)

            the difference between samples in y.
            NOTE: if x is provided, this parameter is ignored

        axis (int)

            the axis of y along which the derivative has to be calculated.

    Output:

        z (ndarray)

            an array being the first derivative of y

    References:

        Winter DA.
            Biomechanics and Motor Control of Human Movement. Fourth Ed.
            Hoboken, New Jersey: John Wiley & Sons Inc; 2009.
    '''
    # validate the input data
    _validate_arr(y, 1, type=[float, int])

    # get x
    if x is None:
        _validate_obj(dt, int)
        x = np.arange(len(y)) * dt
    else:
        _validate_arr(x, shape=y.shape, type=y.dtype)

    # get the derivative
    return (y[2:] - y[:-2]) / (x[2:] - x[:-2])



def d2y(y, x = None, dt = 1):
    '''
    return the second derivative of y.

    Input:

        y (ndarray with shape [n,])

            the signal to be derivated

        x (ndarray with shape [n,] or None)

            the optional signal from which y has to  be derivated

        dt (float)

            the difference between samples in y.
            NOTE: if x is provided, this parameter is ignored

        axis (int)

            the axis of y along which the derivative has to be calculated.

    Output:

        z (ndarray)

            an array being the second derivative of y

    References:

        Winter DA.
            Biomechanics and Motor Control of Human Movement. Fourth Ed.
            Hoboken, New Jersey: John Wiley & Sons Inc; 2009.
    '''
    # validate the input data
    _validate_arr(y, 1, type=[float, int])

    # get x
    if x is None:
        _validate_obj(dt, int)
        x = np.arange(len(y)) * dt
    else:
        _validate_arr(x, shape=y.shape, type=y.dtype)

    # get the derivative
    dy = (y[2:] - y[1:-1]) / (x[2:] - x[1:-1])
    dy -= (y[1:-1] - y[:-2]) / (x[1:-1] - x[:-2])
    dx = (x[2:] - x[:-2]) * 0.5
    return dy / dx



def pad(y, before=0, after=0, value=0):
    """
    pad the signal at its ends.

    Input:

        y (1D array)

            the signal to be padded

        before (int)

            the number of padding values to be added before x.

        after (int)

            the number of padding values to be added after x.

        value (float)

            the value to be used for padding.

    Output:

        z (1D array)

            the padded array
    """
    # validate the inputs
    _validate_arr(y)
    _validate_obj(after, int)
    _validate_obj(before, int)
    _validate_obj(value, float)

    # get the pads
    a_pad = np.tile(value, after)
    b_pad = np.tile(value, before)

    # concatenate the signal
    return np.concatenate([b_pad, y, a_pad], axis=0).flatten()



def rescale_arr(y, vmin = 0, vmax = 1):
    """
    scale the array y to have minimum equal to vmin and maximum equal to vmax.

    Input:

        y (1D array)

            the signal to be scaled

        vmin (float)

            the minimum value allowed by the rescaled signal.

        vmax (int)

            the maximum value allowed by the rescaled signal.

    Output:

        z (1D array)

            the rescaled array
    """

    # validate the data
    _validate_arr(y)
    _validate_obj(vmin, (float, int))
    _validate_obj(vmax, (int, float))
    assert vmin <= vmax, "'vmin' must be <= 'vmax'."

    # rescale
    z = (y - np.min(y)) / (np.max(y) - np.min(y))
    return z * (vmax - vmin) + vmin



def freedman_diaconis_bins(y):
    """
    return a digitized version of y where each value is linked to a
    bin (i.e an int value) according to the rule.

                             IQR(x)
            width = 2 * ---------------
                        len(x) ** (1/3)

    Input:

        y (1D array)

            a ndarray that has to be digitized.

    Output:

        d (1D array)

            an array with the same shape of y containing the index
            of the bin of which the corresponding sample of y is part.

    References:

        Freedman D, Diaconis P.
            (1981) On the histogram as a density estimator:L 2 theory.
            Z. Wahrscheinlichkeitstheorie verw Gebiete 57: 453–476.
            doi: 10.1007/BF01025868
    """

    # validate
    _validate_arr(y)

    # y IQR
    q1 = np.quantile(y, 0.25)
    q3 = np.quantile(y, 0.75)
    iqr = q3 - q1

    # get the width
    w = 2 * iqr / (len(y) ** (1 / 3))

    # get the number of intervals
    n = int(np.floor(1 / w)) + 1

    # digitize z
    d = np.zeros(y.shape)
    for i in (np.arange(n) + 1):
        d[np.argwhere((y >= (i - 1) * w) & (y < i * w)).flatten()] = i - 1
    return d



def mean_filter(y, n=1, offset=0.5):
    """
    mean filter.

    Input:

        y (1D array)

            a 1D array signal to be filtered.

        n (int)

            the order of the filter.

        offset (float)

            should the window be centered around the current sample?.
            The offset value should range between -1 and 1 where:

            0   means that at each i-th sample the window of length n
                over which the mean is calculated has the i-th sample as
                starting value.

            1   means that for the same i-th sample, n-1 samples
                behind i will be used to calculate the mean. An offset of

            0.5 will consider i centered within the filtering window.

    Output:

        z (1D array)

            The filtered signal.
    """

    # control inputs
    _validate_arr(y)
    _validate_obj(n, (int))
    _validate_obj(offset, (float, int))
    txt = "'offset' must be a float in the [-1, 1] range."
    assert offset >= -1, txt
    assert offset <= 1, txt

    # get the window range
    w = np.unique(np.int(rescale_arr(np.arange(n), 0, 1) - offset * n))

    # get the indices of the samples
    i = [w + j for j in np.arange(len(y))]

    # validate the indices (i.e. remove negative values and those above len(y))
    i = [j[np.argwhere((j >= 0) & (j < len(y))).flatten()] for j in i]

    # get the mean of each window
    return np.array([np.mean(y[j]) for j in i]).flatten()



def median_filter(y, n=1, offset=0.5):
    """
    median filter.

    Input:

        y (1D array)

            a 1D array signal to be filtered.

        n (int)

            the order of the filter.

        offset (float)

            should the window be centered around the current sample?.
            The offset value should range between -1 and 1 where:

            0   means that at each i-th sample the window of length n
                over which the mean is calculated has the i-th sample as
                starting value.

            1   means that for the same i-th sample, n-1 samples
                behind i will be used to calculate the mean. An offset of

            0.5 will consider i centered within the filtering window.

    Output:

        z (1D array)

            The filtered signal.
    """

    # control inputs
    _validate_arr(y)
    _validate_obj(n, (int))
    _validate_obj(offset, (float, int))
    txt = "'offset' must be a float in the [-1, 1] range."
    assert offset >= -1, txt
    assert offset <= 1, txt

    # get the window range
    w = np.unique(np.int(rescale_arr(np.arange(n), 0, 1) - offset * n))

    # get the indices of the samples
    i = [w + j for j in np.arange(len(y))]

    # validate the indices (i.e. remove negative values and those above len(y))
    i = [j[np.argwhere((j >= 0) & (j < len(y))).flatten()] for j in i]

    # get the median of each window
    return np.array([np.median(y[j]) for j in i]).flatten()



def interpolate_cs(y, n=None, x_old=None, x_new=None):
    """
    Get the cubic spline interpolation of y.

    Input:

        y (1D array)

            the data to be interpolated.

        n (int)

            the number of points for the interpolation.

        x_old (1D array)

            the x coordinates corresponding to y.
            It is ignored if n is provided.

        x_new (1D array)

            the newly (interpolated) x coordinates corresponding to y.
            It is ignored if n is provided.

    Output:

        z (1D array)

            the interpolated y axis
    """

    # control of the inputs
    _validate_arr(y)
    if n is not None:
        _validate_obj(n, (int))
        x_old = np.arange(len(y))
        x_new = np.linspace(np.min(x_old), np.max(x_old), n)
    else:
        _validate_arr(x_old, shape=y.shape)
        _validate_arr(x_new, dim=y.ndim)

    # get the cubic-spline interpolated y
    cs = si.CubicSpline(x_old, y)
    return cs(x_new)



def residuals_analysis(y, fs, f_num=1000, f_max=None, segments=2,
                       min_samples=2, which_segment=None, filt_fun=None,
                       filt_opt=None):
    """
    Perform Winter's residual analysis of y.

    Input:

        x (1D array)

            the signal to be investigated

        fs (float)

            the sampling frequency of the signal.

        f_num (int)

            the number of frequencies to be tested within the (0, f_max)
            range to create the residuals curve of the Winter's residuals
            analysis approach.

        f_max (float)

            the maximum filter frequency that is tested. If None, it is
            defined as the frequency covering the 99% of the cumulative
            signal power.

        segments (int)

            the number of segments that can be used to fit the residuals
            curve in order to identify the best deflection point.
            NOTE: values above 3 will greatly increase the computation time.

        min_samples (int)

            the minimum number of elements that have to be considered for
            each segment during the calculation of the best deflection point.

        which_segment (int or None)

            the segment to be considered as the one allowing the calculation
            of the optimal cut-off. It must be an int in the [1, segments]
            range. If None, the segment resulting in the most flat line from
            those that have been calculated is used.

        filt_fun (function)

            the filter to be used for the analysis. If None, a Butterworth,
            low-pass, 4th order phase-corrected filter is used.
            If a function is provided, two arguments are mandatory:

              - the data, passed as first argument
              - the cutoffs, passed as second argument.

        filt_opt (dict)

            the options for the filter. If not None, a dict containing the
            key-values combinations to be passed to filt_fun.

    Output:

        cutoff (float)

            the suggested cutoff value

        SSEs (pandas.DataFrame)

            a pandas.DataFrame with the selected frequencies as index and the
            Sum of Squared Residuals as columns.

    Procedure:

        the signal is filtered over a range of frequencies and the sum of
        squared residuals (SSE) against the original signal is computer for
        each tested cut-off frequency. Next, a series of fitting lines are
        used to estimate the optimal disruption point defining the cut-off
        frequency optimally discriminating between noise and good quality
        signal.

    References:

        Winter DA. Biomechanics and Motor Control of Human Movement.
            Fourth Ed. Hoboken, New Jersey: John Wiley & Sons Inc; 2009.

        Lerman PM. Fitting Segmented Regression Models by Grid Search.
            Appl Stat. 1980;29(1):77.
    """

    # control the inputs
    _validate_arr(y)
    _validate_obj(fs, float)
    assert fs > 0, "'fs' must be > 0."
    _validate_obj(f_num, int)
    assert f_num > 1, "'f_num' must be > 1."
    if f_max is None:
        P, F = psd(y, fs)
        f_max = np.arghwere(np.cumsum(P) / np.sum(P) >= 0.99).flatten()
        f_max = np.min([fs / 2, F[f_max[0]]])
    else:
        _validate_obj(f_max, (float))
    _validate_obj(segments, int)
    _validate_obj(min_samples, int)
    assert min_samples >= 2, "'min_samples' must be >= 2."
    if which_segment is not None:
        _validate_obj(which_segment, int)
        txt = "'which_segment' must be an int in the [1, {}] range."
        txt = txt.format(segments)
        assert which_segment > 1, txt
    if filt_fun is None:
        filt_fun = butt_filt
    _validate_obj(filt_opt, dict)
    if filt_opt is None:
        filt_opt = {
            'order': 4,
            'fs': fs,
            'type': 'lowpass',
            'phase_corrected': True
            }
    else:
        _validate_obj(filt_opt, dict)

    # get the frequency span
    freqs = np.linspace(0, f_max, f_num + 1)[1:]

    # get the SSEs
    Q = [np.sum((y - filt_fun(y, i, **filt_opt)) ** 2) for i in freqs]
    Q = np.array(Q)

    # reshape the SSE as dataframe
    D = pd.DataFrame(Q, index=freqs, columns=['SSE'])

    # get the optimal crossing over point that separates the S regression
    # lines best fitting the residuals data.
    F = crossovers(Q, segments, min_samples)[1]

    # get the intercept of optimal line
    # if which_segment is None, find the fitting line being the most flat
    if which_segment is None:
        I = F[np.argmin([abs(i[0]) for i in F])][1]
    else:
        I = F[which_segment][1]

    # get the optimal cutoff
    opt = freqs[np.argmin(abs(Q - I))]

    # return the parameters
    return opt, D



def crossovers(y, segments=2, min_samples=5):
    """
    Detect the position of the crossing over points between K regression
    lines used to best fit the data.

    Input:

        y (1D array)

            The data to be fitted.

        segments (int)

            the number of segments that can be used to fit the residuals
            curve in order to identify the best deflection point.
            NOTE: values above 3 will greatly increase the computation time.

        min_samples (int)

            the minimum number of elements that have to be considered for
            each segment during the calculation of the best deflection point.

    Output:

        C (1D array)

            An ordered array of indices containing the samples corresponding
            to the detected crossing over points.

        F (list)

            a list of tuples containing the slope and intercept of the line
            describing each fitting segment.

    Procedure:

        1)  Get all the segments combinations made possible by the given
            number of crossover points.

        2)  For each combination, calculate the regression lines corresponding
            to each segment.

        3)  For each segment calculate the residuals between the calculated
            regression line and the effective data.

        5)  Once the sum of the residuals have been calculated for each
            combination, sort them by residuals amplitude.

    References:

        Lerman PM.
            Fitting Segmented Regression Models by Grid Search.
            Appl Stat. 1980;29(1):77.
    """

    # control the inputs
    _validate_arr(y)
    _validate_obj(segments, int)
    _validate_obj(min_samples, int)
    assert min_samples >= 2, "'min_samples' must be >= 2."

    # get the residuals calculating formula
    def SSEs(x, y, s):

        # get the coordinates
        C = [np.arange(s[i], s[i + 1] + 1) for i in np.arange(len(s) - 1)]

        # get the fitting parameters for each interval
        Z = [np.polyfit(x[i], y[i], 1) for i in C]

        # get the regression lines for each interval
        V = [np.polyval(v, x[C[i]]) for i, v in enumerate(Z)]

        # get the sum of squared residuals
        return np.sum([np.sum((y[C[i]] - v) ** 2) for i, v in enumerate(V)])

    # get the X axis
    x = np.arange(len(y))

    # get all the possible combinations of segments
    J = [np.arange(min_samples * i, len(y) - min_samples * (segments - i))
         for i in np.arange(1, segments)]
    J = [j for j in it.product(*J)]

    # remove those combinations having segments shorter than "samples"
    J = [i for i in J if np.all(np.diff(i) >= min_samples)]

    # generate the crossovers matrix
    J = np.hstack((
        np.zeros((len(J), 1)),
        np.atleast_2d(J),
        np.ones((len(J), 1)) * len(y) - 1
        )).astype(int)

    # calculate the residuals for each combination
    R = np.array([SSEs(x, y, i) for i in J])

    # sort the residuals
    T = np.argsort(R)

    # get the optimal crossovers order
    O = x[J[T[0]]]

    # get the fitting slopes
    F = [np.arange(i0, i1) for i0, i1 in zip(O[:-1], O[1:])]
    F = [np.polyfit(i, y[i], 1) for i in F]

    # return the crossovers
    return O[1:-1], F



def butt_filt(y, cutoff, fs, order=4, type='lowpass', phase_corrected=True):
    """
    Provides a convenient function to call a Butterworth filter with the
    specified parameters.

    Input:

        y (1D array)

            the signal to be filtered.

        cutoff (float, list, ndarray)

            the filter cutoff in Hz.

        fs (float)

            the sampling frequency of the signal in Hz.

        type (str)

            a string defining the type of the filter
            e.g. "low", "high", "bandpass", etc.

        phase_corrected (bool)

            should the filter be applied twice in opposite directions to
            correct for phase lag?

    Output:

        z (1D array)

            the resulting 1D filtered signal
    """

    # control the inputs
    _validate_arr(y)
    _validate_obj(cutoff, (float, list, np.ndarray))
    if isinstance(cutoff, (np.ndarray, list)):
        assert len(cutoff) == 2, "'cutoff' length must be 2."
        txt = "all cutoff values must be float or int"
        assert np.all([isinstance(i, (float, int)) for i in cutoff]), txt
    _validate_obj(order, int)
    _validate_obj(type, str)
    _validate_obj(phase_corrected, bool)

    # get the filter coefficients
    sos = ss.butter(
        order,
        (np.array([cutoff]).flatten() / (0.5 * fs)),
        type,
        output="sos"
        )

    # get the filtered data
    return ss.sosfiltfilt(sos, y) if phase_corrected else ss.sosfilt(sos, y)



def psd(y, fs=1, n=None):
    """
    compute the power spectrum of y using fft

    Input:

        y (1D array)

            A 1D numpy array

        fs (float)

            the sampling frequency

        n (None, int)

            the number of samples to be used for FFT.
            If None, the length of y is used.

    Output:

        P (1D ndarray)

            the power of each frequency

        F (1D ndarray)

            the frequencies.
    """
    # control the inputs
    _validate_arr(y)
    _validate_obj(fs, (int, float))

    # set n
    if n is None:
        n = len(y)
    else:
        _validate_obj(n, int)

    # get the FFT and normalize by the length of y
    f = np.fft.rfft(y - np.mean(y), n) / len(y)

    # get the amplitude of the signal
    a = abs(f)

    # get the power of the signal
    P = np.concatenate([[a[0]], 2 * a[1:-1], [a[-1]]]).flatten() ** 2

    # get the frequencies
    F = np.linspace(0, fs / 2, len(P))

    # return the data
    return P, F



def find_peaks(y, height=None):
    """
    detect the location (in sample units) of the peaks within y.

    Input:

        y (1D array)

            a 1D signal

        height (float, None)

            a scalar defining the minimum height to be considered
            for a valid peak. If None, all peaks are returned.

    Output:

        p (1D array)

            the indices of the peaks (in sample units) in y.
    """
    # control the inputs
    _validate_arr(y)
    if height is None:
        height = np.min(y)
    else:
        _validate_obj(height, (float))

    # get the first derivative of the signal
    d1 = d1y(y)

    # get the sign of the first derivative
    d1[d1 == 0] = 1
    sn = d1 / abs(d1)

    # get the peaks
    zc = np.argwhere(sn[1:] - sn[:-1] == -2).flatten()

    # exclude all peaks below the set height
    return zc[y[zc] >= height] + 1



def crossings(y, value=0.):
    """
    Dectect the crossing points in x compared to value.

    Input:

        y (1D array)

            the 1D signal from which the crossings have to be found.

        value (float or 1D array with the same shape of y)

            the value/s to be used to detect the crossings. If value
            is an array, the function will find those points in x
            crossing values according to the value of values at a
            given location.

    Output:

        c (1D array)

            the samples corresponding to the crossings.

        s (ndarray)

            the sign of the crossings. Positive sign means crossings
            where the signal moves from values lower than "value" to
            values higher than "value". Negative sign indicate the
            opposite trend.
    """
    # control of the inputs
    _validate_arr(y)
    _validate_obj(value, float)

    # get the sign of the signal without the offset
    sn = y - value
    sn[sn == 0.] = 1
    sn = sn / abs(sn)

    # get the location of the crossings
    cr = np.argwhere(abs(sn[1:] - sn[:-1]) == 2).flatten()

    # return the crossings
    return sn, cr



def xcorr(y, biased=False, full=False, *args):
    """
    set the (multiple) auto/cross correlation of the data in y.

    Input:

        y (1D array)

            the signal from which the auto or cross-correlation is provided.

        biased (bool)

            if True, the biased auto/cross-correlation is provided.
            Otherwise, the 'unbiased' estimator is returned.

        full (bool)

            Should the negative lags be reported?

        *args

            additional signals to be used for cross-correlation.

    Output:

        x (1D array)

            the auto/cross-correlation value.

        l (1D array)

            the lags in sample units.
    """
    # control the inputs
    _validate_arr(y)
    _validate_obj(biased, bool)
    _validate_obj(full, bool)
    for i in args:
        _validate_arr(args[i])

    # take the autocorrelation if only y is provided
    if len(args) == 0:
        Y = np.atleast_2d(y)
        X = np.vstack((Y, Y))

    # otherwise calculate the cross-correlation
    else:
        Y = np.atleast_2d(y)
        A = np.vstack([np.atleast_2d(a) for a in args])
        X = np.vstack([Y, A])

    # get the matrix shape
    P, N = X.shape

    # remove the mean from each dimension
    V = X - np.atleast_2d(np.mean(X, 1)).T

    # take the cross-correlation
    xc = []
    for i in np.arange(P - 1):
        for j in np.arange(i + 1, P):
            R = np.atleast_2d(ss.fftconvolve(V[i], V[j][::-1], "full"))
            xc = np.vstack((xc, R)) if len(xc) > 0 else np.copy(R)

    # average over all the multiples
    xc = np.mean(xc, 0)

    # adjust the output
    lags = np.arange(-(N - 1), N)
    if not full:
        xc = xc[(N - 1):]
        lags = lags[(N - 1):]

    # normalize
    xc /= (N + 1 - abs(lags)) if not biased else (N + 1)

    # return the cross-correlation data
    return xc, lags



def magnitude(y, base=10):
    """
    return the order in the given base of the value

    Input:

        value (float)

            the value to be checked

        base (float)

            the base to be used to define the order of the number

    Output:

        mag (float)

            the number required to elevate the base to get the value
    """
    _validate_obj(y, (float, int))
    _validate_obj(base, (float, int))

    # return the magnitude
    if y == 0 or base == 0:
        return 0
    else:
        return np.log(abs(y)) / np.log(base)



def get_files(path, extension='', check_subfolders=False):
    """
    list all the files having the required extension in the
    provided folder and its subfolders (if required).

    Input:

        path (str)

            a directory where to look for the files.

        extension (str)

            a str object defining the ending of the files that have
            to be listed.

        check_subfolders (bool)

            if True, also the subfolders found in path are searched,
            otherwise only path is checked.

    Output:

        files (list)
            a list containing the full_path to all the files corresponding
            to the input criteria.
    """
    # control the inputs
    _validate_obj(path, str)
    _validate_obj(extension, str)
    _validate_obj(check_subfolders, bool)

    # output storer
    out = []

    # surf the path by the os. walk function
    for root, files in os.walk(path)[0, 2]:
        for obj in files:
            if obj[-len(extension):] == extension:
                out += [os.path.join(root, obj)]

        # handle the subfolders
        if not check_subfolders:
            break

    # return the output
    return out



def to_excel(path, df, sheet="Sheet1", keep_index=True, new_file=False):
    """
    a shorthand function to save a pandas dataframe to an excel file

    Input:

        path (str)

            the path to the file where to store the file.

        df (pandas.DataFrame)

            a pandas.DataFrame.

        sheet (str)

            the sheet name.

        keep_index (boolean)

            if True, the dataframe index is preserved.
            Otherwise it is ignored.

        new_file (boolean)

            if True, a completely new file will be created.

    Output:

        The data stored to the indicated file.
    """

    # control the inputs
    _validate_obj(path, str)
    _validate_obj(df, pd.DataFrame)
    _validate_obj(sheet, str)
    _validate_obj(keep_index, bool)
    _validate_obj(new_file, bool)

    # get the workbook
    if os.path.exists(path) and not new_file:
        wb = xl.load_workbook(path)
    else:
        wb = xl.Workbook()
        try:
            sh = wb["Sheet"]
            wb.remove(sh)
        except Exception:
            pass

    # get the sheet
    try:
        sh = wb[sheet]
        wb.remove(sh)
    except Exception:
        pass
    sh = wb.create_sheet(sheet)

    # write the headers
    [R, C] = df.shape
    if keep_index:
        index = np.atleast_2d(df.index.tolist())
        if index.shape[0] == 1:
            index = index.T
        data_cols = index.shape[1] + 1
    else:
        data_cols = 1
    header = np.atleast_2d(df.columns.tolist())
    if header.shape[0] == 1:
        header = header.T
    data_rows = header.shape[1] + 1
    for i, col in enumerate(header):
        for j, el in enumerate(col):
            ch = data_cols + i
            rh = 1 + j
            sh.cell(rh, ch, el)
    if keep_index:
        for i, row in enumerate(index):
            for j, el in enumerate(row):
                ri = data_rows + i
                ci = 1 + j
                sh.cell(ri, ci, el)

    # write the data
    V = df.values
    for r in range(R):
        for c in range(C):
            sh.cell(data_rows + r, data_cols + c, V[r, c])

    # save data
    wb.save(path)



def from_excel(path, sheets=None, **kwargs):
    """
    a shorthand function to collect data from an excel file
    and to store them into a dict.

    Input:

        path (str)

            the path to the file where to store the file.

        sheets (list of str)

            the name of the sheets to be imported. If None
            all sheets will be imported.

        kwargs

            additional arguments passed to pandas.read_excel

    Output:

        a dict object with keys equal to the sheets name and pandas
        dataframe as elements of each sheet in the excel file.
    """

    # retrive the data in the path file
    try:
        xlfile = pd.ExcelFile(path)
        sheets = np.array(xlfile.sheet_names
                          if sheets is None else [sheets]).flatten()
    except Exception:
        sheets = []
    finally:
        xlfile.close()

    # return the dict
    return {i: pd.read_excel(path, i, **kwargs) for i in sheets}



def get_time(tic=None, toc=None, as_string=True, compact=True):
    """
    get the days, hours, minutes and seconds between the two times.
    If only tic is provided, it is considered as the lapsed time.
    If neither tic nor toc are provided, the function returns the
    current time as float.

    Input (optional)

        tic (int)

            an integer representing the starting time

        toc (int)

            an integer indicating the stopping time

        as_string (bool)

            should the output be returned as string?

        compact (bool)

            if "as_string" is true, should the time be reported in a
            compact or in an extensive way?

    Output:

        If nothing is provided, the function returns the current time.
        If only tic is provided, the function returns the time value
        from it to now. If both tic and toc are provided, the function
        returns the time difference between them.
    """

    # check what to do
    if tic is None:
        return time.time()
    elif toc is None:
        tm = np.float(tic)
    else:
        tm = np.float(toc - tic)

    # convert the time value in days, hours, minutes,
    # seconds and milliseconds
    d = int(np.floor(tm / 86400))
    tm -= (d * 86400)
    h = int(np.floor(tm / 3600))
    tm -= (h * 3600)
    m = int(np.floor(tm / 60))
    tm -= (m * 60)
    s = int(np.floor(tm))
    tm -= s
    ms = int(np.round(1000 * tm, 0))

    # report the calculated time
    if not as_string:
        return {
            "Days": [d],
            "Hours": [h],
            "Minutes": [m],
            "Seconds": [s],
            "Milliseconds": [ms]
            }
    else:
        st = "{:0>2d}".format(d) + (" Days - " if not compact else ":")
        st += "{:0>2d}".format(h)
        st += " Hours - " if not compact else ":"
        st += "{:0>2d}".format(m)
        st += " Minutes - " if not compact else ":"
        st += "{:0>2d}".format(s)
        st += " Seconds - " if not compact else ":"
        st += "{:0>3d}".format(ms)
        st += " Milliseconds" if not compact else ""
        return st



def lvlup(path):
    """
    Goes to the superior level in the directory path.

    Input:

        path (str)

            a file or a directory. Otherwise a message is casted and
            the function returns the input as is.

    Output:

        s (str)

            a string reflecting the superior directory of file.
    """

    # control the inputs
    _validate_obj(path, str)

    # return the upper level
    return os.path.sep.join(path.split(os.path.sep)[:-1])



def nan_replace_SVR(y, support=[], max_tr_data=1000,
                    GridSearchKwargs={}, SVRKwargs={}):
    '''
    Use Support Vector Regression (SVR) to provide the coordinates of the
    missing samples in the current vector.

    Input:

        support (iterable of 1D arrays)

            the list of arrays whoose values can be
            used as features to train the SVR model.

        max_tr_data (int)

            the maximum number of training data to be used.
            If the effective available number is lower than max_tr_data,
            all the training data are used. Otherwise, the specified number
            is randomly sampled from the available pool.

        SVRKwargs (dict)

            default = {
                "kernel": "rbf",
                "gamma": "scale",
                "tol": 1e-5,
                "epsilon": 5e-4,  # i.e. 0.5 mm error.
                "max_iter": 1e4
                }

            parameters passed to the SVR class. Full documentation can be
            found here:
            https://scikit-learn.org/stable/modules/generated/sklearn.svm.SVR.html

        GridSearchKwargs (dict)

            default = {
                "estimator": SVR(**SVRKwargs),
                "param_grid": {
                    "C": np.unique([i * (10 ** j)
                                    for i in np.arange(1, 11)
                                    for j in np.linspace(-10, -1, 10)])
                    },
                "scoring": "neg_mean_absolute_error"
                }

            parameters passed to the scikit-learn GridSearchCV class.
            Full documentation can be found here:
            https://scikit-learn.org/stable/modules/generated/sklearn.model_selection.GridSearchCV.html#sklearn.model_selection.GridSearchCV

    Output:

        z (1D array)

            a 1D array with the same shape of y without missing values.

    References:

        Smola A. J., Schölkopf B. (2004).
            A tutorial on support vector regression.
            Statistics and Computing, 14(3), 199–222.
    '''

    # control of the inputs
    _validate_arr(y)
    _validate_obj(support, list)
    for s in support:
        _validate_arr(s, shape=y.shape)
    _validate_obj(max_tr_data, int)
    _validate_obj(SVRKwargs, dict)
    _validate_obj(GridSearchKwargs, dict)

    # default SVR estimator options
    opt_SVRKwargs = {
        "kernel": "rbf",
        "gamma": "scale",
        "tol": 1e-5,
        "epsilon": 5e-4,  # i.e. 0.5 mm error.
        "max_iter": 1e4
        }

    # update the SVR options with those provided by the user
    opt_SVRKwargs.update(**SVRKwargs)

    # default GridSearchCV options
    opt_GridSearchKwargs = {
        "estimator": SVR(**opt_SVRKwargs),
        "param_grid": {
            "C": np.unique([i * (10 ** j) for i in np.arange(1, 11)
                            for j in np.linspace(-10, -1, 10)])
            },
        "scoring": "neg_mean_absolute_error"
        }

    # update the GridSearchCV options with those provided by the user
    opt_GridSearchKwargs.update(**GridSearchKwargs)

    # get a copy of the current vector
    complete = np.copy(y)

    # get the missing values
    miss_idx = np.argwhere(np.isnan(y)).flatten()

    # replace missing data
    if len(miss_idx) > 0 and len(support) > 0:

        # get the training dataset
        x = np.vstack([np.atleast_2d(v) for v in support]).T

        # exclude the sets containing missing data from the training sets
        tr_idx = np.arange(x.shape[0])
        tr_idx = tr_idx[~miss_idx]
        tr_idx = tr_idx[np.any(~np.isnan(x[tr_idx, :]), axis=1).flatten()]

        # get max_tr_data unique samples at random
        tr_idx = np.random.permutation(tr_idx)[:max_tr_data]
        training_set = x[tr_idx, :]
        if training_set.shape[0] > 0:

            # grid searcher
            grid = GridSearchCV(**opt_GridSearchKwargs)
            est = grid.fit(training_set, y[tr_idx])

            # replace the missing data
            rep = est.best_estimator_.predict(x[miss_idx, :])
            complete[miss_idx] = rep

        return complete



def nan_replace(y, value=None):
    """
    replace missing values in y.


        Input

            y (1D array)

                the data which contains missing values

            value (None, float)

                the value to be used to replace the data.
                If None, cubic spline interpolation is used to cover the
                missing values.
                If float value is provided, all missing data are
                replaced with it.

        Output:

        z (1D array)

            a 1D array with the same shape of y without missing values.
    """

    # control of the inputs
    _validate_arr(y)
    z = np.copy(y)
    miss_idx = np.argwhere(np.isnan(y)).flatten()

    # find the proper replacing value
    if value is not None:
        _validate_obj(value, (float, int))
        z[miss_idx] = value

    # use cubic spline interpolation
    else:

        # valid data
        train_idx = np.arghwere(~np.isnan(y)).flatten()

        # obtain the cubic spline interpolated data
        z = interpolate_cs(
            y = y[train_idx],
            x_old = train_idx,
            x_new = np.arange(len(y))
            )

    # return the interpolated data
    return z
